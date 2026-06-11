"""
export.py — write a formatted .xlsx of the ranked screen.

Used by both the CLI (writes a file) and the Streamlit app (returns bytes for a
download button). Formatting: frozen header, bold header fill, %/number formats,
and a green color-scale on the composite column.
"""

import io

import pandas as pd

import config

PCT_COLS = ["implied_upside", "pos_52w", "dist_sma200", "rev_growth", "roe", "ann_vol"]
PRICE_COLS = ["price", "target_mean", "target_high", "target_low"]
NUM2_COLS = ["fwd_pe", "pe", "ps", "ev_ebitda", "beta", "rsi", "composite"]

DISPLAY = ["name", "sector", "price", "composite", "buy_zone", "implied_upside",
           "target_mean", "rsi", "pos_52w", "dist_sma200", "fwd_pe", "rev_growth",
           "roe", "beta", "ann_vol", "score_analyst", "score_buy_zone",
           "score_valuation", "score_quality", "score_momentum"]


def to_excel_bytes(ranked: pd.DataFrame, sheet: str = "Screen") -> bytes:
    """Return a formatted .xlsx as bytes."""
    cols = [c for c in DISPLAY if c in ranked.columns]
    df = ranked[cols].reset_index()  # ticker becomes a column

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        df.to_excel(xl, index=False, sheet_name=sheet)
        wb, ws = xl.book, xl.sheets[sheet]
        _format(ws, df)
    return buf.getvalue()


def _format(ws, df) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    cur = config.CURRENCY

    # Header styling + freeze
    head_fill = PatternFill("solid", fgColor="1B5E20")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    colmap = {name: i + 1 for i, name in enumerate(df.columns)}
    nrows = len(df) + 1

    def fmt(colname, number_format):
        if colname in colmap:
            letter = get_column_letter(colmap[colname])
            for r in range(2, nrows + 1):
                ws[f"{letter}{r}"].number_format = number_format

    for c in PCT_COLS:
        fmt(c, "0.0%")
    for c in PRICE_COLS:
        fmt(c, f'"{cur}"#,##0.00')
    for c in NUM2_COLS:
        fmt(c, "0.0")

    # Color scale on composite
    if "composite" in colmap:
        letter = get_column_letter(colmap["composite"])
        rng = f"{letter}2:{letter}{nrows}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max", end_color="2E7D32"))

    # Column widths
    for i, name in enumerate(df.columns, start=1):
        width = 22 if name == "name" else (12 if name not in ("sector",) else 18)
        ws.column_dimensions[get_column_letter(i)].width = width
